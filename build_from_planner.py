#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Inyecta el export de Excel de Microsoft Planner en planner-kpi.html como
"corte" (snapshot) embebido, para que el dashboard muestre los datos reales
sin necesidad de conexión.

Uso:
    python3 build_from_planner.py "ruta/al/PLANNER_PERIFERIA_URBANA.xlsx"

Cómo obtener el Excel: Planner (en Teams) → el plan → "..." → Exportar plan
a Excel. Vuelva a ejecutar el script con cada nuevo export para refrescar
el corte. La conexión en vivo (botón "Conectar con Microsoft") sigue
funcionando y tiene prioridad sobre el corte.

Requiere: pip install openpyxl
"""
import json
import sys
from pathlib import Path

import openpyxl

HTML = Path(__file__).parent / "planner-kpi.html"
MARK_START = "/*__SNAPSHOT_START__*/"
MARK_END = "/*__SNAPSHOT_END__*/"


def day(value):
    """Normaliza fechas del export ('YYYY-MM-DD', datetime o vacío) a 'YYYY-MM-DD'."""
    if value is None or value == "":
        return None
    s = str(value)[:10]
    return s if len(s) == 10 and s[4] == "-" else None


def split_users(value):
    return [u.strip() for u in str(value or "").split(";") if u.strip()]


def read_export(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    meta = next(wb["Planificar"].iter_rows(min_row=2, max_row=2, values_only=True))
    plan_name, exported = str(meta[1] or "Planner"), day(meta[2])

    # "Datos consolidados" trae nombres legibles (en "Tareas" son IDs)
    ws = wb["Datos consolidados"]
    header = [str(c or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    col = {name: idx for idx, name in enumerate(header)}
    ixs = {
        "titulo": col["Nombre de la tarea"],
        "bucket": col["Depósito"],
        "estado": col["Estado"],
        "asignado": col["Asignado a"],
        "vence": col["Fecha de vencimiento"],
        "fin": col["Fecha de finalización"],
    }

    tasks = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[ixs["titulo"]]:
            continue
        completado = str(row[ixs["estado"]] or "").strip() == "Completado"
        tasks.append({
            "t": str(row[ixs["titulo"]]).strip(),
            "d": day(row[ixs["vence"]]),
            # solo cuenta como completada si el estado lo dice (la fecha sola no basta)
            "f": day(row[ixs["fin"]]) if completado else None,
            "u": split_users(row[ixs["asignado"]]),
            "b": str(row[ixs["bucket"]] or "").strip(),
        })
    return {"plan": plan_name, "exported": exported, "tasks": tasks}


def inject(snapshot):
    html = HTML.read_text(encoding="utf-8")
    start = html.index(MARK_START) + len(MARK_START)
    end = html.index(MARK_END)
    payload = json.dumps(snapshot, ensure_ascii=False, separators=(",", ":"))
    # </script> dentro de un título rompería el bloque <script>
    payload = payload.replace("</", "<\\/")
    html = html[:start] + "\nconst SNAPSHOT = " + payload + ";\n" + html[end:]
    HTML.write_text(html, encoding="utf-8")


def main():
    if len(sys.argv) != 2:
        sys.exit(__doc__)
    snapshot = read_export(sys.argv[1])
    inject(snapshot)
    done = sum(1 for t in snapshot["tasks"] if t["f"])
    print(f"OK: {len(snapshot['tasks'])} tareas ({done} completadas) del plan "
          f"\"{snapshot['plan']}\" · corte {snapshot['exported']} → {HTML.name}")


if __name__ == "__main__":
    main()
