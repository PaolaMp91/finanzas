#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_dashboard.py
==================
Genera el dashboard PMO de PLANIFICACIÓN (Fase 2 · Boulevard Sur) a partir del
archivo de "Seguimiento Semanal" más reciente de Teams/SharePoint.

Uso:
    # Con un archivo local (export de MS Project a Excel):
    python3 build_dashboard.py "13 07 2026 DASHBOARD PLANIFICACION.xlsx"

    # Sin archivo -> usa los datos verificados de semilla (SEED) y produce
    # igual el HTML (se completará al 100% cuando el motor de refresco lea el
    # archivo real vía Microsoft Graph). Ver refresh_from_teams.py.
    python3 build_dashboard.py

Salida: dashboard.html  (un solo archivo, listo para publicar y pegar en Teams).

El Excel se espera con columnas tipo MS Project (los nombres se detectan de
forma flexible, sin importar mayúsculas/espacios/acentos):
    Name | Outline_Level | Percent_Complete | Start | Finish | Duration |
    Resource_Names | Predecessors
Las filas de resumen (Outline_Level 1 y 2) llevan el % de avance acumulado que
MS Project calcula automáticamente por rama.
"""
import sys, os, json, re, unicodedata, datetime

# --------------------------------------------------------------------------- #
#  DATOS VERIFICADOS DE SEMILLA (cuando no hay Excel disponible localmente)
#  Extraídos del archivo "13 07 2026 DASHBOARD PLANIFICACION.xlsx".
#  El motor de refresco (Graph API) sobreescribe todo esto con el archivo real.
# --------------------------------------------------------------------------- #
SEED = {
    "corte": "13/07/2026",
    "archivo": "13 07 2026 DASHBOARD PLANIFICACION.xlsx",
    "carpeta": "TECNICO-BOULEVARSUR › PLANIFICACION › 01. PLANNER › FASE 2 › 02. Seguimiento Semanal",
    "sincronizado": False,
    "proyecto": {
        "nombre": "PROGRAMA DE PLANIFICACIÓN FASE 2 · BOULEVARD SUR",
        "inicio": "2025-10-29",
        "fin": "2027-04-09",
        "duracion_dias": 378,
        "tareas": 1109,
        "avance": 0.03,          # ~3% (control de avance 13/06 y 18/06); se sincroniza al valor exacto
    },
    # Fases de nivel 2 con datos verificados. avance=None => se completa al sincronizar.
    "fases": [
        {"nombre": "DISEÑO Y PLANIFICACIÓN FASE 2", "inicio": "2025-10-29", "fin": "2027-01-18", "duracion_dias": 319, "avance": None, "responsable": "MARS / PERIFERIA"},
        {"nombre": "ANTEPROYECTO FASE 2",           "inicio": "2025-10-29", "fin": "2026-08-05", "duracion_dias": 201, "avance": None, "responsable": "MARS / PERIFERIA"},
        {"nombre": "URBANIZACIÓN FASE 2",           "inicio": "2026-06-11", "fin": "2027-01-18", "duracion_dias": 158, "avance": None, "responsable": "MARS / Especialistas"},
        {"nombre": "TORRE D",                        "inicio": None, "fin": None, "duracion_dias": None, "avance": None, "responsable": "MARS"},
        {"nombre": "TORRE E",                        "inicio": None, "fin": None, "duracion_dias": None, "avance": None, "responsable": "MARS"},
        {"nombre": "TORRE F",                        "inicio": None, "fin": None, "duracion_dias": None, "avance": None, "responsable": "MARS"},
        {"nombre": "TORRE P-2",                      "inicio": None, "fin": None, "duracion_dias": None, "avance": None, "responsable": "MARS"},
        {"nombre": "AMENIDADES F-2",                 "inicio": None, "fin": None, "duracion_dias": None, "avance": None, "responsable": "MARS"},
        {"nombre": "TRÁMITES (MSPAS · DGAC · MARN · BANCO/FHA)", "inicio": None, "fin": None, "duracion_dias": None, "avance": None, "responsable": "PERIFERIA"},
    ],
    # Distribución por responsable (conteo de tareas). Vacío => se completa al sincronizar.
    "responsables": {},
    "responsables_lista": ["PERIFERIA", "MARS", "GEOESTUDIOS", "WASYMA", "INSEL", "INCCO", "NABLA", "FHA", "MSPAS", "DGAC", "MARN", "MYMA"],
    # Próximos hitos (verificados de la porción inicial del archivo).
    "hitos": [
        {"nombre": "Aprobación de Anteproyecto Fase 2", "fecha": "2026-06-18", "responsable": "MARS / PERIFERIA"},
        {"nombre": "Inicio de Planificación con Especialistas", "fecha": "2026-06-19", "responsable": "MARS / PERIFERIA"},
        {"nombre": "Validación de estudio de suelos", "fecha": "2026-07-22", "responsable": "GEOESTUDIOS"},
        {"nombre": "Primera Homologación de planificación", "fecha": "2026-08-05", "responsable": "MARS / PERIFERIA"},
        {"nombre": "Definición de plataformas y movimiento de tierras", "fecha": "2026-10-06", "responsable": "MARS / PERIFERIA"},
    ],
}

# --------------------------------------------------------------------------- #
#  Lectura del Excel (MS Project export)
# --------------------------------------------------------------------------- #
def _norm(s):
    """normaliza encabezados: minúsculas, sin acentos, sin separadores."""
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode()
    return re.sub(r"[^a-z0-9]", "", s.lower())

def _to_date(v):
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    if isinstance(v, str):
        for fmt in ("%m/%d/%Y %H:%M", "%m/%d/%Y", "%Y-%m-%d", "%d/%m/%Y", "%d/%m/%Y %H:%M"):
            try:
                return datetime.datetime.strptime(v.strip(), fmt).date()
            except ValueError:
                pass
    return None

def _pct(v):
    """acepta 0-1, 0-100 o '35%' -> fracción 0-1."""
    if v is None or v == "":
        return None
    if isinstance(v, str):
        v = v.strip().replace("%", "")
        try:
            v = float(v)
        except ValueError:
            return None
    v = float(v)
    return v / 100.0 if v > 1.0 else v

def _dur_days(v):
    if v is None:
        return None
    m = re.search(r"(\d+(?:[.,]\d+)?)", str(v))
    return int(round(float(m.group(1).replace(",", ".")))) if m else None

def extract_from_xlsx(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Hoja vacía")

    # localizar la fila de encabezados (la que contiene 'Name'/'Nombre')
    hdr_idx, hdr = 0, rows[0]
    for i, r in enumerate(rows[:10]):
        cells = [_norm(c) for c in r if c is not None]
        if any(c in ("name", "nombre", "nombredetarea") for c in cells):
            hdr_idx, hdr = i, r
            break
    H = {_norm(c): j for j, c in enumerate(hdr) if c is not None}

    def find(*names):
        for n in names:
            if n in H:
                return H[n]
        # coincidencia parcial
        for n in names:
            for k, j in H.items():
                if n in k:
                    return j
        return None

    ci = {
        "name":  find("name", "nombre", "nombredetarea"),
        "lvl":   find("outlinelevel", "nivel", "niveldeesquema"),
        "pct":   find("percentcomplete", "completado", "completo", "avance"),
        "start": find("start", "comienzo", "inicio"),
        "finish":find("finish", "fin"),
        "dur":   find("duration", "duracion"),
        "res":   find("resourcenames", "recursos", "responsable"),
    }
    data = rows[hdr_idx + 1:]

    def cell(row, key):
        j = ci[key]
        return row[j] if (j is not None and j < len(row)) else None

    today = datetime.date.today()

    # ---- 1ª pasada: recoger TODAS las tareas en una lista ---------------------
    T = []
    for row in data:
        name = cell(row, "name")
        if not name or not str(name).strip():
            continue
        lvl = cell(row, "lvl")
        try:
            lvl = int(lvl) if lvl is not None else None
        except (ValueError, TypeError):
            lvl = None
        res = cell(row, "res")
        res = str(res).strip() if res else ""
        T.append({
            "name": str(name).strip(), "lvl": lvl,
            "pct": _pct(cell(row, "pct")),
            "d0": _to_date(cell(row, "start")),
            "d1": _to_date(cell(row, "finish")),
            "dur": _dur_days(cell(row, "dur")),
            "res": res,
        })
    n = len(T)

    def is_leaf(i):
        """hoja = la siguiente tarea NO es hija (nivel <= al actual)."""
        if i == n - 1:
            return True
        a, b = T[i]["lvl"], T[i + 1]["lvl"]
        if a is None or b is None:
            return bool(T[i]["res"])   # respaldo: con recurso se asume hoja
        return b <= a

    def rollup_res(i0, i1):
        """responsables distintos de las hojas dentro del rango [i0, i1)."""
        seen = []
        for k in range(i0, i1):
            if is_leaf(k) and T[k]["res"]:
                for r in re.split(r"[,;/]", T[k]["res"]):
                    r = r.strip()
                    if r and r not in seen:
                        seen.append(r)
        return seen

    # ---- conteos por responsable (solo hojas, para no duplicar resúmenes) -----
    responsables = {}
    sin_resp = 0
    for i, t in enumerate(T):
        if not is_leaf(i):
            continue
        if t["res"]:
            for r in re.split(r"[,;/]", t["res"]):
                r = r.strip()
                if r:
                    responsables[r] = responsables.get(r, 0) + 1
        else:
            sin_resp += 1

    # ---- proyecto (nivel 1) y fases (nivel 2) con responsable heredado --------
    proj = None
    fases = []
    for i, t in enumerate(T):
        if t["lvl"] == 1 and proj is None:
            resp = rollup_res(i, n)
            proj = {"nombre": t["name"], "inicio": t["d0"].isoformat() if t["d0"] else None,
                    "fin": t["d1"].isoformat() if t["d1"] else None,
                    "duracion_dias": t["dur"], "avance": t["pct"]}
        elif t["lvl"] == 2:
            # rango de la fase: hasta la siguiente tarea de nivel <= 2
            j = i + 1
            while j < n and (T[j]["lvl"] is None or T[j]["lvl"] > 2):
                j += 1
            resp = rollup_res(i, j) or ([t["res"]] if t["res"] else [])
            fases.append({
                "nombre": t["name"],
                "inicio": t["d0"].isoformat() if t["d0"] else None,
                "fin": t["d1"].isoformat() if t["d1"] else None,
                "duracion_dias": t["dur"], "avance": t["pct"],
                "responsable": " · ".join(resp[:3]) if resp else "Sin asignar",
            })

    # ---- tareas ATRASADAS: hojas con fin < hoy y avance < 100% ----------------
    atrasadas = []
    for i, t in enumerate(T):
        if not is_leaf(i):
            continue
        pct = t["pct"] or 0.0
        if t["d1"] and t["d1"] < today and pct < 0.999:
            atrasadas.append({
                "nombre": t["name"],
                "responsable": t["res"] or "⚠ Sin asignar",
                "fin": t["d1"].isoformat(),
                "avance": pct,
                "dias": (today - t["d1"]).days,
            })
    atrasadas.sort(key=lambda a: a["dias"], reverse=True)

    # ---- hitos: tareas de 0-1 día a futuro con nombre de gate -----------------
    hitos = []
    for i, t in enumerate(T):
        if t["dur"] is not None and t["dur"] <= 1 and t["d1"] and t["d1"] >= today and \
           re.search(r"homolog|aprob|entrega|inicio|validac|revision|desembolso", t["name"], re.I):
            hitos.append({"nombre": t["name"], "fecha": t["d1"].isoformat(),
                          "responsable": t["res"] or "Sin asignar"})
    hitos = sorted(hitos, key=lambda h: h["fecha"])[:8]

    if proj is None:
        proj = {"nombre": "PROGRAMA DE PLANIFICACIÓN FASE 2", "inicio": None,
                "fin": None, "duracion_dias": None, "avance": None}
    proj["tareas"] = n

    # fecha de corte a partir del nombre del archivo
    fname = os.path.basename(path)
    m = re.search(r"(\d{1,2})[ _\-]+(\d{1,2})[ _\-]+(\d{4})", fname)
    corte = f"{int(m.group(1)):02d}/{int(m.group(2)):02d}/{m.group(3)}" if m else \
            today.strftime("%d/%m/%Y")

    return {
        "corte": corte, "archivo": fname, "sincronizado": True,
        "carpeta": SEED["carpeta"],
        "proyecto": proj, "fases": fases,
        "responsables": responsables,
        "responsables_lista": sorted(responsables) or SEED["responsables_lista"],
        "sin_responsable": sin_resp,
        "atrasadas": atrasadas,
        "hitos": hitos,
    }

# --------------------------------------------------------------------------- #
#  Métricas derivadas (avance de calendario, días, etc.)
# --------------------------------------------------------------------------- #
def enrich(data):
    p = data["proyecto"]
    today = datetime.date.today()
    s = _to_date(p.get("inicio")); e = _to_date(p.get("fin"))
    cal = {"span": None, "elapsed": None, "restante": None, "pct_cal": None, "hoy_pos": None}
    if s and e and e > s:
        span = (e - s).days
        elapsed = max(0, min(span, (today - s).days))
        cal = {
            "span": span, "elapsed": elapsed, "restante": span - elapsed,
            "pct_cal": elapsed / span,
            "hoy_pos": max(0.0, min(1.0, (today - s).days / span)),
            "inicio": s.isoformat(), "fin": e.isoformat(),
        }
    def esperado(a, b):
        """avance esperado por calendario (línea base) entre fechas a y b."""
        if not a or not b or b <= a:
            return None
        if today <= a:
            return 0.0
        if today >= b:
            return 1.0
        return (today - a).days / (b - a).days

    # avance esperado del proyecto (línea base global)
    p["esperado"] = esperado(s, e)
    if p.get("avance") is not None and p.get("esperado") is not None:
        p["desvio"] = p["avance"] - p["esperado"]

    # posiciones de barras + esperado/desvío por fase
    for f in data["fases"]:
        f["pos"] = None
        fs = _to_date(f.get("inicio")); fe = _to_date(f.get("fin"))
        if s and e and e > s and fs and fe:
            span = (e - s).days
            f["pos"] = {"left": max(0.0, (fs - s).days / span),
                        "width": max(0.012, (fe - fs).days / span)}
        f["esperado"] = esperado(fs, fe)
        f["desvio"] = (f["avance"] - f["esperado"]) \
            if (f.get("avance") is not None and f.get("esperado") is not None) else None

    data["calendario"] = cal
    data["hoy"] = today.strftime("%d/%m/%Y")
    data["generado"] = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
    data.setdefault("atrasadas", [])
    data.setdefault("sin_responsable", 0)
    return data

# --------------------------------------------------------------------------- #
#  Render HTML (un solo archivo)
# --------------------------------------------------------------------------- #
def render(data):
    return TEMPLATE.replace("/*__DATA__*/", "const D = " + json.dumps(data, ensure_ascii=False) + ";")

TEMPLATE = r"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1.0"/>
<meta http-equiv="refresh" content="3600"/>
<title>PMO · Planificación Fase 2 · Boulevard Sur</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js"></script>
<style>
  :root{--navy:#0f2c4d;--navy2:#16406e;--blue:#1f73c4;--sky:#e8f0fa;--gold:#c9a227;--teal:#0f9d8c;
    --bg:#eef1f6;--card:#fff;--line:#e2e8f0;--txt:#1c2733;--muted:#67748a;--soft:#8a96a8;
    --good:#1f9d57;--bad:#d64545;--warn:#d98a1f;--shadow:0 1px 3px rgba(16,35,60,.07),0 6px 18px rgba(16,35,60,.05);}
  *{box-sizing:border-box;margin:0;padding:0}
  body{font-family:"Segoe UI",system-ui,-apple-system,Arial,sans-serif;background:var(--bg);color:var(--txt);line-height:1.45;font-size:13.5px}
  .wrap{max-width:1300px;margin:0 auto;padding:0 22px 55px}
  .brand{background:linear-gradient(120deg,var(--navy),var(--navy2) 55%,#1d4f86);color:#fff;padding:24px 0;box-shadow:var(--shadow)}
  .brand .wrap{display:flex;flex-wrap:wrap;justify-content:space-between;align-items:flex-end;gap:16px;padding-bottom:0}
  .brand .kick{color:var(--gold);font-size:11px;font-weight:800;letter-spacing:2.5px;text-transform:uppercase}
  .brand h1{font-size:23px;font-weight:700;margin-top:4px}
  .brand .sub{color:#b9cce4;font-size:12.5px;margin-top:5px}
  .brand .meta{text-align:right;font-size:12px;color:#cdddf0}
  .live{background:rgba(31,157,87,.22);border:1px solid rgba(82,217,135,.5);color:#aef0c8;font-size:11px;font-weight:700;padding:4px 11px;border-radius:20px;display:inline-flex;align-items:center;gap:6px}
  .live.warn{background:rgba(217,138,31,.18);border-color:rgba(217,138,31,.5);color:#f2cf95}
  .dot{width:7px;height:7px;border-radius:50%;background:#52d987;display:inline-block;box-shadow:0 0 0 3px rgba(82,217,135,.25)}
  .live.warn .dot{background:var(--gold);box-shadow:0 0 0 3px rgba(201,162,39,.25)}
  .corte{display:inline-block;background:rgba(255,255,255,.14);border:1px solid rgba(255,255,255,.22);padding:5px 12px;border-radius:20px;font-weight:600;color:#fff;margin-top:8px}
  section{margin-top:28px}
  h2.sec{font-size:12px;text-transform:uppercase;letter-spacing:1.1px;color:var(--navy);font-weight:800;margin-bottom:14px;display:flex;align-items:center;gap:9px}
  h2.sec::before{content:"";width:4px;height:15px;background:var(--gold);border-radius:3px}
  h2.sec .hint{font-weight:600;letter-spacing:0;text-transform:none;color:var(--soft);font-size:11.5px;margin-left:auto}
  .grid{display:grid;gap:14px}.g6{grid-template-columns:repeat(6,1fr)}.g4{grid-template-columns:repeat(4,1fr)}
  .g3{grid-template-columns:repeat(3,1fr)}.g2{grid-template-columns:1.35fr 1fr}
  @media(max-width:1040px){.g6{grid-template-columns:repeat(3,1fr)}.g4{grid-template-columns:repeat(2,1fr)}.g3{grid-template-columns:1fr}.g2{grid-template-columns:1fr}}
  @media(max-width:560px){.g6,.g4{grid-template-columns:repeat(2,1fr)}}
  .card{background:var(--card);border:1px solid var(--line);border-radius:12px;padding:16px;box-shadow:var(--shadow)}
  .kpi .kl{color:var(--muted);font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:.4px}
  .kpi .kv{font-size:24px;font-weight:800;color:var(--navy);margin-top:6px;line-height:1.1}
  .kpi .ks{color:var(--soft);font-size:11.5px;margin-top:5px}
  .kpi.hero{background:linear-gradient(150deg,var(--navy),var(--navy2));border:none}
  .kpi.hero .kl{color:#a9c2e0}.kpi.hero .kv{color:#fff}.kpi.hero .ks{color:#94aece}
  .kpi.gold .kv{color:var(--gold)}.kpi.teal .kv{color:var(--teal)}.kpi.warn .kv{color:var(--warn)}.kpi.bad .kv{color:var(--bad)}
  .dev{font-weight:700}.dev.neg{color:var(--bad)}.dev.pos{color:var(--good)}
  .atr td .rp{display:inline-block;font-size:10.5px;font-weight:700;color:#9a3d16;background:#fdeee6;border:1px solid #f4cbb4;padding:2px 8px;border-radius:20px}
  .atr td .rp.no{color:var(--bad);background:#fbe9e9;border-color:#f2c4c4}
  .atr .dd{color:var(--bad);font-weight:800}
  /* progreso circular */
  .gauge{display:flex;align-items:center;gap:16px}
  .ring{--v:0;--c:var(--gold);width:92px;height:92px;border-radius:50%;flex:0 0 auto;
    background:conic-gradient(var(--c) calc(var(--v)*1%),#e6ecf4 0);display:flex;align-items:center;justify-content:center;position:relative}
  .ring::before{content:"";position:absolute;inset:9px;background:var(--card);border-radius:50%}
  .ring b{position:relative;font-size:20px;font-weight:800;color:var(--navy)}
  .ring.onnavy::before{background:var(--navy)}.ring.onnavy b{color:#fff}
  /* barra */
  .bar{height:9px;border-radius:6px;background:#e6ecf4;overflow:hidden;margin-top:7px}
  .bar > i{display:block;height:100%;border-radius:6px;background:linear-gradient(90deg,var(--blue),var(--teal))}
  /* tabla fases */
  table{width:100%;border-collapse:collapse;font-size:13px}
  th,td{text-align:right;padding:10px 11px;border-bottom:1px solid var(--line);vertical-align:middle}
  th:first-child,td:first-child{text-align:left}
  thead th{color:var(--muted);font-size:10.5px;text-transform:uppercase;letter-spacing:.5px;font-weight:700;background:#f7f9fc}
  tbody tr:hover{background:#f6f9fd}
  .ph{font-weight:650;color:var(--navy)}
  .resp{display:inline-block;font-size:10.5px;font-weight:700;color:var(--blue);background:var(--sky);border:1px solid #cfe0f2;padding:2px 8px;border-radius:20px}
  .minibar{width:120px;height:8px;border-radius:5px;background:#e6ecf4;display:inline-block;overflow:hidden;vertical-align:middle;margin-right:8px}
  .minibar > i{display:block;height:100%;background:linear-gradient(90deg,var(--blue),var(--teal))}
  .pend{color:var(--soft);font-style:italic;font-size:11.5px}
  /* timeline / gantt */
  .tl{position:relative;margin-top:6px}
  .tlrow{display:grid;grid-template-columns:210px 1fr;gap:12px;align-items:center;padding:7px 0}
  .tlname{font-size:12.5px;color:var(--navy);font-weight:600;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
  .tltrack{position:relative;height:22px;background:#eef2f7;border-radius:6px}
  .tlbar{position:absolute;top:3px;bottom:3px;border-radius:5px;background:linear-gradient(90deg,var(--navy2),var(--blue));box-shadow:inset 0 0 0 1px rgba(255,255,255,.15)}
  .tltoday{position:absolute;top:0;bottom:0;width:2px;background:var(--gold);z-index:3}
  .tltoday.lbl{top:-6px}
  .tltoday.lbl::after{content:"HOY";position:absolute;top:-16px;left:50%;transform:translateX(-50%);font-size:9px;font-weight:800;color:var(--gold);letter-spacing:.5px}
  .tlaxis{display:grid;grid-template-columns:210px 1fr;gap:12px;margin-top:8px}
  .tlaxis .ax{display:flex;justify-content:space-between;color:var(--soft);font-size:10.5px}
  .tlpend{position:absolute;inset:3px 0;border-radius:5px;border:1.5px dashed #c3ceda;background:repeating-linear-gradient(45deg,#f3f6fa,#f3f6fa 6px,#eaeff5 6px,#eaeff5 12px)}
  /* hitos */
  .hito{display:flex;gap:12px;align-items:flex-start;padding:11px 0;border-bottom:1px dashed var(--line)}
  .hito:last-child{border:none}
  .hdate{flex:0 0 auto;text-align:center;background:var(--sky);border:1px solid #cfe0f2;border-radius:9px;padding:6px 9px;min-width:56px}
  .hdate .d{font-size:17px;font-weight:800;color:var(--navy);line-height:1}
  .hdate .m{font-size:10px;font-weight:700;color:var(--blue);text-transform:uppercase;letter-spacing:.5px}
  .hbody .hn{font-weight:600;color:var(--txt)}
  .hbody .hr{color:var(--soft);font-size:11.5px;margin-top:2px}
  .hdias{margin-left:auto;font-size:11px;font-weight:700;color:var(--warn);white-space:nowrap}
  .chips{display:flex;flex-wrap:wrap;gap:7px;margin-top:4px}
  .chip{background:var(--sky);color:var(--blue);font-size:11.5px;font-weight:600;padding:5px 11px;border-radius:8px;border:1px solid #cfe0f2}
  .chart-box{position:relative;height:250px}
  .note{color:var(--soft);font-size:11.5px;margin-top:12px}
  footer{color:var(--muted);font-size:11.5px;border-top:1px solid var(--line);padding-top:18px;margin-top:34px}
  footer code{background:#eef2f7;padding:1px 6px;border-radius:4px;color:var(--navy);font-size:11px}
  footer b{color:var(--navy)}
  .banner{background:#fff7e6;border:1px solid #f0d79a;color:#8a6d1f;border-radius:10px;padding:11px 14px;font-size:12px;margin-top:14px;display:flex;gap:9px;align-items:flex-start}
  .banner b{color:#6b520f}
</style>
</head>
<body>
  <header class="brand">
    <div class="wrap">
      <div>
        <div class="kick">Periferia Urbana · PMO</div>
        <h1>Planificación Fase 2 · Boulevard Sur</h1>
        <div class="sub" id="projName">—</div>
        <span class="corte" id="corte">Corte —</span>
      </div>
      <div class="meta">
        <div id="liveBadge"></div>
        <div style="margin-top:9px" id="folder">—</div>
        <div style="margin-top:3px">Actualizado <b id="gen">—</b></div>
      </div>
    </div>
  </header>

  <div class="wrap">
    <div id="syncBanner"></div>

    <!-- KPIs -->
    <section>
      <h2 class="sec">Resumen del programa</h2>
      <div class="grid g6" id="kpis"></div>
    </section>

    <!-- Avance calendario vs ejecución -->
    <section>
      <h2 class="sec">Avance <span class="hint">Calendario transcurrido vs. ejecución de tareas</span></h2>
      <div class="grid g2">
        <div class="card">
          <div class="gauge">
            <div class="ring onnavy" id="ringCal" style="--c:var(--gold)"><b id="ringCalTxt">—</b></div>
            <div>
              <div class="kpi"><div class="kl">Calendario transcurrido</div>
              <div class="kv" id="calDias">—</div>
              <div class="ks" id="calSub">—</div></div>
            </div>
          </div>
          <div style="margin-top:14px">
            <div class="gauge">
              <div class="ring" id="ringEjec" style="--c:var(--teal)"><b id="ringEjecTxt">—</b></div>
              <div>
                <div class="kpi"><div class="kl">Ejecución de tareas</div>
                <div class="kv teal" id="ejecTxt">—</div>
                <div class="ks" id="ejecSub">Avance físico del plan (% MS Project)</div></div>
              </div>
            </div>
          </div>
          <div class="note" id="gapNote"></div>
        </div>
        <div class="card">
          <div class="kl" style="color:var(--muted);font-weight:700;font-size:11px;text-transform:uppercase;letter-spacing:.4px">Distribución de tareas por responsable</div>
          <div class="chart-box" style="height:220px"><canvas id="respChart"></canvas></div>
          <div id="respChips" class="chips"></div>
        </div>
      </div>
    </section>

    <!-- Cronograma -->
    <section>
      <h2 class="sec">Cronograma por fase <span class="hint">Ventana del programa · marcador HOY</span></h2>
      <div class="card">
        <div class="tl" id="timeline"></div>
        <div class="tlaxis"><div></div><div class="ax"><span id="axStart">—</span><span id="axEnd">—</span></div></div>
      </div>
    </section>

    <!-- Tareas atrasadas -->
    <section>
      <h2 class="sec">Tareas atrasadas <span class="hint" id="atrHint">—</span></h2>
      <div class="card" style="padding:0;overflow:hidden">
        <table class="atr">
          <thead><tr><th>Tarea</th><th>Responsable</th><th>Fin planeado</th><th>Avance</th><th>Días de atraso</th></tr></thead>
          <tbody id="atrBody"></tbody>
        </table>
      </div>
      <div class="note" id="atrNote"></div>
    </section>

    <!-- Tabla de fases -->
    <section>
      <h2 class="sec">Fases de planificación <span class="hint">Avance real vs. esperado por calendario (línea base)</span></h2>
      <div class="card" style="padding:0;overflow:hidden">
        <table>
          <thead><tr><th>Fase</th><th>Responsable</th><th>Inicio</th><th>Fin</th><th>Avance</th><th>Esperado</th><th>Desvío</th></tr></thead>
          <tbody id="fasesBody"></tbody>
        </table>
      </div>
    </section>

    <!-- Hitos -->
    <section>
      <h2 class="sec">Próximos hitos y homologaciones</h2>
      <div class="card" id="hitos"></div>
    </section>

    <footer>
      <div>Fuente: <b id="fFolder">—</b> › <code id="fFile">—</code></div>
      <div style="margin-top:7px">Este tablero se regenera automáticamente tomando el archivo de <b>fecha más cercana a hoy</b> dentro de la carpeta de Seguimiento Semanal. La página se refresca sola cada hora.</div>
    </footer>
  </div>

<script>
/*__DATA__*/
(function(){
  const $=id=>document.getElementById(id);
  const pf=v=>v==null?null:Math.round(v*100);
  const fmtD=iso=>{if(!iso)return"—";const p=iso.split("-");return p.length===3?`${p[2]}/${p[1]}/${p[0]}`:iso;};
  const MES=["ENE","FEB","MAR","ABR","MAY","JUN","JUL","AGO","SEP","OCT","NOV","DIC"];

  $("projName").textContent=D.proyecto.nombre||"—";
  $("corte").textContent="Corte "+(D.corte||"—");
  $("folder").textContent=D.archivo||"";
  $("gen").textContent=D.generado||"—";
  $("fFolder").textContent=D.carpeta||"—";
  $("fFile").textContent=D.archivo||"—";
  $("axStart").textContent=fmtD(D.calendario&&D.calendario.inicio);
  $("axEnd").textContent=fmtD(D.calendario&&D.calendario.fin);

  // badge live / sincronización
  if(D.sincronizado){
    $("liveBadge").innerHTML='<span class="live"><span class="dot"></span>Conectado a Teams</span>';
  }else{
    $("liveBadge").innerHTML='<span class="live warn"><span class="dot"></span>Vista previa · pendiente de sincronizar</span>';
    $("syncBanner").innerHTML='<div class="banner"><span>⚙️</span><div><b>Vista previa con datos verificados.</b> El % de avance exacto por fase, las fechas de las torres y la distribución por responsable se completan automáticamente cuando el motor de refresco lee el archivo real de Teams (ver <code>refresh_from_teams.py</code>). La estructura, fechas de fases principales, duración y ventana del proyecto ya son reales.</div></div>';
  }

  // KPIs
  const cal=D.calendario||{};
  const desP=D.proyecto.desvio!=null?pf(D.proyecto.desvio):null;
  const kv=[
    {l:"Duración del plan",v:(D.proyecto.duracion_dias||"—")+" días",s:(D.proyecto.tareas||"—")+" tareas",cls:"hero"},
    {l:"Calendario transcurrido",v:(cal.pct_cal!=null?pf(cal.pct_cal)+"%":"—"),s:(cal.elapsed!=null?cal.elapsed+" de "+cal.span+" días":""),cls:"gold"},
    {l:"Ejecución de tareas",v:(D.proyecto.avance!=null?pf(D.proyecto.avance)+"%":"—"),s:"avance físico real",cls:"teal"},
    {l:"Avance esperado (línea base)",v:(D.proyecto.esperado!=null?pf(D.proyecto.esperado)+"%":"—"),s:"lo que debería llevar hoy",cls:""},
    {l:"Desvío vs línea base",v:(desP!=null?(desP>0?"+":"")+desP+" pts":"—"),s:(desP!=null&&desP<0?"por debajo de lo planeado":"según plan"),cls:"warn"},
    {l:"Tareas atrasadas",v:(D.atrasadas?D.atrasadas.length:0),s:(D.sin_responsable?D.sin_responsable+" sin responsable":"vencidas sin terminar"),cls:"bad"},
  ];
  $("kpis").innerHTML=kv.map(k=>`<div class="card kpi ${k.cls}"><div class="kl">${k.l}</div><div class="kv">${k.v}</div><div class="ks">${k.s}</div></div>`).join("");

  // anillos
  const calP=cal.pct_cal!=null?pf(cal.pct_cal):0;
  $("ringCal").style.setProperty("--v",calP);$("ringCalTxt").textContent=calP+"%";
  $("calDias").textContent=(cal.elapsed!=null?cal.elapsed:"—")+" días";
  $("calSub").textContent=(cal.restante!=null?cal.restante+" días restantes":"");
  const ejP=D.proyecto.avance!=null?pf(D.proyecto.avance):null;
  $("ringEjec").style.setProperty("--v",ejP||0);$("ringEjecTxt").textContent=ejP!=null?ejP+"%":"—";
  $("ejecTxt").textContent=ejP!=null?ejP+"%":"—";
  if(ejP!=null&&calP!=null){
    const gap=calP-ejP;
    $("gapNote").innerHTML=gap>15
      ? `⚠️ El calendario avanza <b style="color:var(--warn)">${gap} pts</b> por delante de la ejecución física. Riesgo de holgura consumida.`
      : `El calendario y la ejecución están alineados (brecha ${gap} pts).`;
  }

  // responsables chart (blindado: si Chart.js no carga, muestra chips y NO rompe el resto)
  const R=D.responsables||{};const ent=Object.entries(R).sort((a,b)=>b[1]-a[1]).slice(0,10);
  function respChips(){
    const src=ent.length?ent.map(e=>`${e[0]} (${e[1]})`):(D.responsables_lista||[]);
    $("respChart").parentElement.innerHTML=ent.length?"":'<div class="pend" style="padding:20px 4px">Se completa al sincronizar con Teams.</div>';
    $("respChips").innerHTML=src.map(r=>`<span class="chip">${r}</span>`).join("");
  }
  try{
    if(ent.length && typeof Chart!=="undefined"){
      new Chart($("respChart"),{type:"doughnut",data:{labels:ent.map(e=>e[0]),
        datasets:[{data:ent.map(e=>e[1]),backgroundColor:["#16406e","#1f73c4","#0f9d8c","#c9a227","#67748a","#4a90d9","#2bb3a3","#e0b83a","#8a96a8","#0f2c4d"],borderWidth:2,borderColor:"#fff"}]},
        options:{plugins:{legend:{position:"right",labels:{boxWidth:11,font:{size:11}}}},cutout:"58%"}});
    }else{ respChips(); }
  }catch(e){ respChips(); }

  // timeline (etiqueta HOY solo en la primera fila)
  const tl=$("timeline");let html="";
  D.fases.forEach((f,i)=>{
    let bar;
    if(f.pos){bar=`<div class="tlbar" style="left:${(f.pos.left*100).toFixed(1)}%;width:${(f.pos.width*100).toFixed(1)}%"></div>`;}
    else{bar=`<div class="tlpend"></div>`;}
    const today=cal.hoy_pos!=null?`<div class="tltoday${i===0?' lbl':''}" style="left:${(cal.hoy_pos*100).toFixed(1)}%"></div>`:"";
    html+=`<div class="tlrow"><div class="tlname" title="${f.nombre}">${f.nombre}</div><div class="tltrack">${bar}${today}</div></div>`;
  });
  tl.innerHTML=html;

  // tareas atrasadas
  const atr=D.atrasadas||[];
  $("atrHint").textContent=atr.length?`${atr.length} vencidas y sin terminar`:"ninguna 🎉";
  $("atrBody").innerHTML=atr.slice(0,25).map(a=>{
    const p=pf(a.avance);
    const noResp=/sin asignar/i.test(a.responsable);
    return `<tr><td class="ph">${a.nombre}</td><td><span class="rp${noResp?' no':''}">${a.responsable}</span></td><td>${fmtD(a.fin)}</td><td>${p}%</td><td class="dd">${a.dias} d</td></tr>`;
  }).join("")||'<tr><td colspan="5" style="text-align:center;color:var(--good);padding:16px">Sin tareas atrasadas.</td></tr>';
  const notes=[];
  if(atr.length>25) notes.push(`Mostrando las 25 más atrasadas de ${atr.length}.`);
  if(D.sin_responsable) notes.push(`${D.sin_responsable} tarea${D.sin_responsable==1?"":"s"} del plan ${D.sin_responsable==1?"no tiene":"no tienen"} responsable asignado.`);
  $("atrNote").textContent=notes.join("  ·  ");

  // tabla fases (avance real vs esperado / línea base)
  $("fasesBody").innerHTML=D.fases.map(f=>{
    const a=f.avance!=null?pf(f.avance):null;
    const av=a!=null?`<span class="minibar"><i style="width:${a}%"></i></span><b>${a}%</b>`:`<span class="pend">sincroniza</span>`;
    const esp=f.esperado!=null?pf(f.esperado)+"%":"—";
    let dev="—";
    if(f.desvio!=null){const d=pf(f.desvio);dev=`<span class="dev ${d<0?'neg':'pos'}">${d>0?'+':''}${d} pts</span>`;}
    return `<tr><td class="ph">${f.nombre}</td><td>${f.responsable?`<span class="resp">${f.responsable}</span>`:'<span class="resp">Sin asignar</span>'}</td><td>${fmtD(f.inicio)}</td><td>${fmtD(f.fin)}</td><td>${av}</td><td>${esp}</td><td>${dev}</td></tr>`;
  }).join("");

  // hitos
  const today=new Date();
  $("hitos").innerHTML=(D.hitos||[]).map(h=>{
    const p=h.fecha.split("-");const dd=p[2],mm=MES[parseInt(p[1],10)-1];
    const dias=Math.round((new Date(h.fecha)-today)/86400000);
    const et=dias<0?"vencido":dias===0?"hoy":`en ${dias} días`;
    return `<div class="hito"><div class="hdate"><div class="d">${dd}</div><div class="m">${mm}</div></div><div class="hbody"><div class="hn">${h.nombre}</div><div class="hr">${h.responsable||""}</div></div><div class="hdias">${et}</div></div>`;
  }).join("")||'<div class="pend">Sin hitos próximos.</div>';
})();
</script>
</body>
</html>"""

# --------------------------------------------------------------------------- #
def main():
    if len(sys.argv) > 1 and os.path.exists(sys.argv[1]):
        data = extract_from_xlsx(sys.argv[1])
        print(f"[ok] datos leídos de: {sys.argv[1]}")
    else:
        if len(sys.argv) > 1:
            print(f"[aviso] no existe '{sys.argv[1]}'. Uso datos de semilla verificados.")
        data = dict(SEED)
    data = enrich(data)
    html = render(data)
    out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "dashboard.html")
    with open(out, "w", encoding="utf-8") as fh:
        fh.write(html)
    p = data["proyecto"]
    print(f"[ok] dashboard.html generado · corte {data['corte']} · "
          f"{p.get('tareas','?')} tareas · {len(data['fases'])} fases · "
          f"sincronizado={data['sincronizado']}")

if __name__ == "__main__":
    main()
